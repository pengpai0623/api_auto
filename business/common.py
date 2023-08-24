import requests
import os
import json
from openpyxl import load_workbook


# 目前我的依赖封装函数和fixture并没有区分很清楚
# 目前来看，接口并不需要区分很清楚
# 但是自动化估计会经常需要
# 或者这里也可以添加三个参数，分别是url parmas 和methon = 0(0为get请求，1为post请求)
def add_topics(parmas):
    url = 'https://cnodejs.org/api/v1/topics'
    res = requests.post(url=url, json=parmas)
    return res


# 获取计划数据
def get_basic_info():
    access_token = '123456'
    base_url = 'https://cnodejs.org/api/v1'
    return base_url, access_token


# 获取测试数据路径
def get_path():
    path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    json_path = r'%s\data\data.json' % path
    csv_path = r'%s\data\data.xlsx' % path
    log_path = r'%s\logs\log_output.log' % path
    return json_path, csv_path, log_path


# 注意，像我们写的这两个获取json和csv格式的函数，其实固定了只能获得指定的文件和sheet页签
# 所以我们就可以把文件路径和页签名称当做参数，这样就更灵活，包括上面的get_path
# 获取json格式数据
def get_json_data():
    path = get_path()[0]
    with open(path, "r", encoding='utf8') as f:
        #  注意json文件格式
        r = f.read()
        test_data1 = json.loads(r)['test_data1']
    return test_data1


# 获取csv格式数据
def get_csv_data():
    path = get_path()[1]
    wb = load_workbook(path)
    # 获取页签
    # print(wb.worksheets)
    # 获取指定页签/指定单元格数据
    ws = wb['data']
    # excel有多少行
    # print(len(tuple(ws.rows)))
    # excel有多少列
    # print(len(tuple(ws.columns)))

    # test_data是获取所有数据
    # testcase_data是获取每一行的数据
    test_data = []
    for x in range(2, len(tuple(ws.rows)) + 1):  # 先遍历第一行
        testcase_data = []
        for y in range(2, len(tuple(ws.columns)) + 1):  # 遍历第一行的所有列
            # 获取第几行第几列的值
            testcase_data.append(ws.cell(row=x, column=y).value)
            # alist.append(testcase_data)
        test_data.append(testcase_data)
    return test_data

